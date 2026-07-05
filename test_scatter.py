import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series

wb = openpyxl.Workbook()
ws_raw = wb.active
ws_raw.title = "原始数据"

ws_raw.cell(row=1, column=1, value="Time")
for i in range(1, 91):
    ws_raw.cell(row=1, column=i+1, value=f"CH{i}")

chart = ScatterChart()
chart.title = "Scatter Test"
chart.style = 13
chart.x_axis.title = "Time"
chart.y_axis.title = "Temp"

xvalues = Reference(ws_raw, min_col=1, min_row=2, max_row=2000)
for i in range(2, 92):
    values = Reference(ws_raw, min_col=i, min_row=1, max_row=2000)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)

ws_raw.add_chart(chart, "A10")
wb.save("test_scatter.xlsx")
print("Done")
