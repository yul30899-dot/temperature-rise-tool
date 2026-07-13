# coding=utf-8
import openpyxl
from openpyxl.chart import LineChart
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
import zipfile
import os
import copy

def process(size):
    temp_file = f'temp_template_{size}.xlsx'
    final_file = f'public/chart_template_compare_{size}.xlsx'
    
    # 1. Use openpyxl to build relationships and layouts
    wb = openpyxl.load_workbook(f'public/chart_template_{size}.xlsx')
    
    ws_compare = wb.create_sheet(title=u"对比原始数据")
    ws_compare.cell(row=1, column=1, value=u"时间")
    for i in range(1, 201):
        ws_compare.cell(row=1, column=i+1, value=f"Compare_CH{i}")
        
    ws_summary = wb['温升数据']
    chart1 = ws_summary._charts[0]
    
    chart2 = LineChart()
    ws_summary.add_chart(chart2)
    
    chart2.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=14, colOff=0, row=1, rowOff=0),
        to=AnchorMarker(col=32, colOff=457200, row=31, rowOff=0)
    )
    chart1.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=14, colOff=0, row=33, rowOff=0),
        to=AnchorMarker(col=32, colOff=457200, row=63, rowOff=0)
    )
    
    wb.save(temp_file)
    
    # 2. Use zipfile to inject original chart XMLs to preserve axes and styling
    with zipfile.ZipFile(f'public/chart_template_{size}.xlsx', 'r') as zorig:
        orig_chart1_xml = zorig.read('xl/charts/chart1.xml').decode('utf-8')
    
    # Chart1: Current Data
    xml1 = orig_chart1_xml.replace(u"<a:t>温升曲线图</a:t>", u"<a:t>温升曲线图 (当前测试数据)</a:t>")
    # Chart2: History Data
    xml2 = orig_chart1_xml.replace(u"<a:t>温升曲线图</a:t>", u"<a:t>温升曲线图 (历史对比数据)</a:t>")
    
    with zipfile.ZipFile(temp_file, 'r') as zin:
        with zipfile.ZipFile(final_file, 'w') as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename == 'xl/charts/chart1.xml':
                    zout.writestr(item, xml1.encode('utf-8'))
                elif item.filename == 'xl/charts/chart2.xml':
                    zout.writestr(item, xml2.encode('utf-8'))
                else:
                    zout.writestr(item, content)
                    
    os.remove(temp_file)
    print(f"Created {final_file}")

if __name__ == "__main__":
    for size in [200, 500, 1000, 2000, 3000, 5000]:
        process(size)
