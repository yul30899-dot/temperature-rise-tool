import openpyxl
from openpyxl.chart import LineChart, Reference

wb = openpyxl.Workbook()

ws_summary = wb.active
ws_summary.title = "温升数据"

ws_raw = wb.create_sheet("原始数据")
ws_raw.cell(row=1, column=1, value="Time")
for i in range(1, 91):
    ws_raw.cell(row=1, column=i+1, value=f"CH{i}")

chart = LineChart()
chart.title = "温升曲线图"
chart.x_axis.title = "时间"
chart.y_axis.title = "温度 (℃)"
chart.width = 40
chart.height = 20

# Put legend at the bottom to avoid squishing the chart
chart.legend.position = "b"

# Define data range (90 channels + time, 2000 rows max)
data = Reference(ws_raw, min_col=2, min_row=1, max_col=91, max_row=2000)
cats = Reference(ws_raw, min_col=1, min_row=2, max_row=2000)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Make lines thinner and smoother for all 90 series
for s in chart.series:
    s.graphicalProperties.line.width = 9525 # 0.75 pt in EMUs (1 pt = 12700 EMUs)
    s.smooth = True # Smooth curves

ws_summary.add_chart(chart, "O2")

wb.save("public/chart_template.xlsx")
print("Template created successfully.")
