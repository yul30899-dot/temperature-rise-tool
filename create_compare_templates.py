import openpyxl
from openpyxl.chart import LineChart, Reference

def create_template(max_rows):
    wb = openpyxl.Workbook()
    
    # Sheet 1: Data 1
    ws_data = wb.active
    ws_data.title = u"\u539f\u59cb\u6570\u636e" # "原始数据"
    
    # Header 1
    ws_data.cell(row=1, column=1, value=u"\u65f6\u95f4") # "时间"
    for i in range(1, 201):
        ws_data.cell(row=1, column=i+1, value=f"CH{i}")
        
    # Sheet 2: Data 2 (Compare)
    ws_compare = wb.create_sheet(title=u"\u5bf9\u6bd4\u539f\u59cb\u6570\u636e") # "对比原始数据"
    ws_compare.cell(row=1, column=1, value=u"\u65f6\u95f4") # "时间"
    for i in range(1, 201):
        ws_compare.cell(row=1, column=i+1, value=f"Compare_CH{i}")

    # Sheet 3: Chart & Summary
    ws_summary = wb.create_sheet(title=u"\u6e29\u5347\u6570\u636e") # "温升数据"
    
    # Chart 1: Primary Data
    chart1 = LineChart()
    chart1.title = u"\u6e29\u5347\u66f2\u7ebf\u56fe (\u5f53\u524d\u6d4b\u8bd5\u6570\u636e)" # "温升曲线图 (当前测试数据)"
    chart1.style = 2
    chart1.y_axis.title = u"\u6e29\u5ea6 (\u2103)" # "温度 (℃)"
    chart1.x_axis.title = u"\u65f6\u95f4" # "时间"
    chart1.width = 30
    chart1.height = 15
    
    data_ref = Reference(ws_data, min_col=2, min_row=1, max_row=max_rows)
    cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=max_rows)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    
    ws_summary.add_chart(chart1, "Q32")
    
    # Chart 2: Compare Data
    chart2 = LineChart()
    chart2.title = u"\u6e29\u5347\u66f2\u7ebf\u56fe (\u5386\u53f2\u5bf9\u6bd4\u6570\u636e)" # "温升曲线图 (历史对比数据)"
    chart2.style = 2
    chart2.y_axis.title = u"\u6e29\u5ea6 (\u2103)" # "温度 (℃)"
    chart2.x_axis.title = u"\u65f6\u95f4" # "时间"
    chart2.width = 30
    chart2.height = 15
    
    data_ref2 = Reference(ws_compare, min_col=2, min_row=1, max_row=max_rows)
    cats_ref2 = Reference(ws_compare, min_col=1, min_row=2, max_row=max_rows)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    
    ws_summary.add_chart(chart2, "Q1")

    filename = f"public/chart_template_compare_{max_rows}.xlsx"
    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    row_counts = [200, 500, 1000, 2000, 3000, 5000]
    for count in row_counts:
        create_template(count)
